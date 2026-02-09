import re
import os
import requests
from fastapi import FastAPI, Request, HTTPException
from openpyxl import load_workbook

# =========================================================
# CONFIG
# =========================================================
SUPPORT_GROUP_ID = int(os.getenv("SUPPORT_GROUP_ID", "-1003575621343"))

# ✅ Claves internas ESTABLES (bot_a/bot_b/bot_c)
BOTS = {
    "bot_a": {
        "token_env": "BOT_TOKEN_A",
        "display": "Ayuda Cajero Referidor",
        "default_error_code": "300",
        "default_error_text": "Error 300, Metabet requiere biométrico",
    },
    "bot_b": {
        "token_env": "BOT_TOKEN_B",
        "display": "Bot Pruebas",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_c": {
        "token_env": "BOT_TOKEN_C",
        "display": "HS Call Center",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_d": {
        "token_env": "BOT_TOKEN_D",
        "display": "TMT",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_e": {
        "token_env": "David 24/7",
        "display": "HS Call Center",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_f": {
        "token_env": "BOT_TOKEN_F",
        "display": "La_Tinca",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_g": {
        "token_env": "BOT_TOKEN_G",
        "display": "Melfast",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_h": {
        "token_env": "BOT_TOKEN_H",
        "display": "Bot Express",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_i": {
        "token_env": "BOT_TOKEN_I",
        "display": "Jhona 24/7",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_j": {
        "token_env": "BOT_TOKEN_J",
        "display": "Team FK",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_k": {
        "token_env": "BOT_TOKEN_K",
        "display": "Nattifast",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_l": {
        "token_env": "BOT_TOKEN_L",
        "display": "Bet Hacks",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_m": {
        "token_env": "BOT_TOKEN_M",
        "display": "Alycor",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_n": {
        "token_env": "BOT_TOKEN_N",
        "display": "BetCajeros",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_o": {
        "token_env": "BOT_TOKEN_O",
        "display": "LA SORTE TEAM",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    "bot_p": {
        "token_env": "BOT_TOKEN_P",
        "display": "EmiFast",
        "default_error_code": "601",
        "default_error_text": "Error 601, la página necesita biométrico",
    },
    
}

# ✅ Alias/nombres que puede mandar C# -> bot_key interno
BOT_ALIASES = {
    "HS Call Center": "bot_c",
    "Ayuda Cajero Referidor": "bot_a",
    "Bot Pruebas": "bot_b",
    # por si te llega en minúsculas o con espacios extra, igual lo cubre normalize_bot_key()
}

# =========================================================
# APP
# =========================================================
app = FastAPI()

CHATID_RE = re.compile(r"ChatID:\s*(-?\d+)")
TICKET_TAG = "🧾 TICKET"

ERROR_CATALOG_PATH = os.getenv("ERROR_CATALOG_PATH", "catalogo_errores.xlsx")
ERROR_MAP = {}  # cache: "604" -> {"plataforma":..., "causa":..., "solucion":...}

SUPPORT_ROUTER_BOT_KEY = os.getenv("SUPPORT_ROUTER_BOT_KEY", "bot_a").strip()
TICKET_API_KEY = os.getenv("TICKET_API_KEY", "").strip()

# =========================================================
# Normalización de bot_key
# =========================================================
def normalize_bot_key(raw: str | None) -> str | None:
    """
    Acepta:
      - bot_key interno: bot_a/bot_b/bot_c
      - alias exacto: 'HS Call Center' -> bot_c
      - display exacto: 'Bot Pruebas' -> bot_b
    Devuelve bot_key interno o None.
    """
    if not raw:
        return None

    s = str(raw).strip()
    if not s:
        return None

    # 1) ya es key interno
    if s in BOTS:
        return s

    # 2) alias exacto
    if s in BOT_ALIASES and BOT_ALIASES[s] in BOTS:
        return BOT_ALIASES[s]

    # 3) alias case-insensitive + trim
    s_low = s.lower()
    for k, v in BOT_ALIASES.items():
        if k.strip().lower() == s_low and v in BOTS:
            return v

    # 4) match por display (case-insensitive)
    for bot_key, bot in BOTS.items():
        if (bot.get("display") or "").strip().lower() == s_low:
            return bot_key

    return None

# =========================================================
# Helpers Telegram
# =========================================================
def get_bot_token(bot_key: str) -> str:
    bot = BOTS.get(bot_key)
    if not bot:
        raise HTTPException(status_code=404, detail="Bot no registrado (bot_key inválido).")

    token = os.getenv(bot["token_env"], "").strip()
    if not token:
        raise RuntimeError(f"Falta variable de entorno {bot['token_env']} para {bot_key}")
    return token

def tg(bot_key: str, method: str, payload: dict):
    token = get_bot_token(bot_key)
    api = f"https://api.telegram.org/bot{token}"
    r = requests.post(f"{api}/{method}", json=payload, timeout=20)
    data = r.json()
    if not data.get("ok"):
        raise RuntimeError(f"Telegram API error calling {method}: {data}")
    return data["result"]

def send_message(
    bot_key: str,
    chat_id: int,
    text: str,
    reply_to_message_id: int | None = None,
    reply_markup: dict | None = None
):
    payload = {"chat_id": chat_id, "text": text}
    if reply_to_message_id:
        payload["reply_to_message_id"] = reply_to_message_id
    if reply_markup:
        payload["reply_markup"] = reply_markup
    return tg(bot_key, "sendMessage", payload)

def answer_callback_query(bot_key: str, callback_query_id: str, text: str = "", show_alert: bool = False):
    payload = {"callback_query_id": callback_query_id, "text": text, "show_alert": show_alert}
    return tg(bot_key, "answerCallbackQuery", payload)

def remove_inline_keyboard(bot_key: str, chat_id: int, message_id: int):
    payload = {
        "chat_id": chat_id,
        "message_id": message_id,
        "reply_markup": {"inline_keyboard": []},
    }
    return tg(bot_key, "editMessageReplyMarkup", payload)

def parse_ticket_botkey(text: str) -> str | None:
    m = re.search(r"BotKey:\s*([^\n\r]+)", text)  # ✅ toma toda la línea (por si trae espacios)
    return m.group(1).strip() if m else None

# =========================================================
# Excel catálogo
# =========================================================
def load_error_catalog(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    out = {}
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        code, plataforma, causa, solucion = row
        if code is None:
            continue

        code_str = str(code).strip()
        if not code_str:
            continue

        out[code_str] = {
            "plataforma": (str(plataforma).strip() if plataforma else "-"),
            "causa": (str(causa).strip() if causa else "-"),
            "solucion": (str(solucion).strip() if solucion else "-"),
        }
    return out

def ensure_error_map_loaded():
    global ERROR_MAP
    if not ERROR_MAP:
        try:
            ERROR_MAP = load_error_catalog(ERROR_CATALOG_PATH)
        except Exception:
            ERROR_MAP = {}

# =========================================================
# Auth para /ticket
# =========================================================
def require_api_key(req: Request):
    if not TICKET_API_KEY:
        return
    key = req.headers.get("x-api-key", "").strip()
    if key != TICKET_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid API key")

# =========================================================
# Health
# =========================================================
@app.get("/")
def health():
    return {"ok": True, "service": "telegram-support-multibot", "bots": list(BOTS.keys())}

# =========================================================
# POST /ticket (desde C#)
# =========================================================
@app.post("/ticket")
async def create_ticket(req: Request):
    require_api_key(req)
    payload = await req.json()

    raw_bot = (payload.get("bot_key") or "").strip()
    bot_key = normalize_bot_key(raw_bot)

    error_code = str(payload.get("error_code") or "").strip()
    user = payload.get("user") or {}

    if not bot_key:
        raise HTTPException(status_code=400, detail=f"bot_key/botname no reconocido: {raw_bot}")
    if not error_code:
        raise HTTPException(status_code=400, detail="error_code requerido")
    if not user.get("id"):
        raise HTTPException(status_code=400, detail="user.id (chatid) requerido")

    client_chat_id = int(user["id"])
    first = (user.get("first_name") or "").strip()
    last = (user.get("last_name") or "").strip()
    username = (user.get("username") or "").strip()
    uname = f"@{username}" if username else "(sin username)"
    full_name = (first + " " + last).strip() or "Cliente"

    bot_display = BOTS[bot_key]["display"]

    
    # =========================================================
    # 🔴 TICKET CHAT (formato fijo, sin catálogo)
    # =========================================================
    if error_code.upper() == "CHAT":
        ticket_text = (
            "🔴TICKET CHAT\n"
            f"👤Cliente: {full_name} {uname}\n"
            "⚠️ El cliente desea comunicarse directamente con alguien de soporte."
            )

        # enviar al grupo usando el router
        send_message(SUPPORT_ROUTER_BOT_KEY, SUPPORT_GROUP_ID, ticket_text)
        return {"ok": True}

    ensure_error_map_loaded()
    info = ERROR_MAP.get(error_code, {"plataforma": "-", "causa": "-", "solucion": "-"})

    # ✅ Guardamos BotKey interno SIEMPRE (para que /r funcione)
    ticket_text = (
        f"{TICKET_TAG}\n"
        f"🤖 Bot: {bot_display}\n"
        f"BotKey: {bot_key}\n"
        f"BotName: {raw_bot}\n"
        f"👤 Cliente: {full_name} {uname}\n"
        f"ChatID: {client_chat_id}\n"
        f"⚠️ Error: Error {error_code}\n"
        f"📝 Plataforma: {info['plataforma']}\n"
        f"🧩 Causa: {info['causa']}\n"
        f"✅ Solución: {info['solucion']}\n\n"
        f"↩️ Responde a ESTE mensaje con:\n"
        f"/r tu respuesta aquí"
    )

    # Ticket al grupo usando el router
    send_message(SUPPORT_ROUTER_BOT_KEY, SUPPORT_GROUP_ID, ticket_text)
    return {"ok": True}

# =========================================================
# Webhook por bot: /telegram/bot_a, /telegram/bot_b, /telegram/bot_c
# =========================================================
@app.post("/telegram/{bot_key}")
async def telegram_webhook(bot_key: str, req: Request):
    if bot_key not in BOTS:
        raise HTTPException(status_code=404, detail="Bot no registrado (bot_key inválido).")

    update = await req.json()
    bot_display = BOTS[bot_key]["display"]

    # 1) CallbackQuery: botón REPORTAR
    if "callback_query" in update:
        cq = update["callback_query"]
        data = cq.get("data", "")
        callback_id = cq.get("id")

        if data.startswith("report:"):
            error_code = data.split(":", 1)[1].strip()

            from_user = cq.get("from", {})
            msg = cq.get("message", {})
            chat = msg.get("chat", {})

            client_chat_id = chat.get("id")
            client_message_id = msg.get("message_id")

            full_name = (from_user.get("first_name", "") + " " + from_user.get("last_name", "")).strip()
            username = from_user.get("username")
            uname = f"@{username}" if username else "(sin username)"

            ensure_error_map_loaded()
            info = ERROR_MAP.get(str(error_code).strip(), {"plataforma": "-", "causa": "-", "solucion": "-"})

            ticket_text = (
                f"{TICKET_TAG}\n"
                f"🤖 Bot: {bot_display}\n"
                f"BotKey: {bot_key}\n"
                f"👤 Cliente: {full_name} {uname}\n"
                f"ChatID: {client_chat_id}\n"
                f"⚠️ Error: Error {error_code}\n"
                f"📝 Plataforma: {info['plataforma']}\n"
                f"🧩 Causa: {info['causa']}\n"
                f"✅ Solución: {info['solucion']}\n\n"
                f"↩️ Responde a ESTE mensaje con:\n"
                f"/r tu respuesta aquí"
            )

            try:
                remove_inline_keyboard(bot_key, client_chat_id, client_message_id)
            except Exception:
                pass

            send_message(bot_key, SUPPORT_GROUP_ID, ticket_text)
            send_message(bot_key, client_chat_id, "✅ Tu reporte fue enviado. En breve soporte se comunicará contigo.")
            answer_callback_query(bot_key, callback_id, "Reporte enviado ✅", show_alert=False)

        return {"ok": True}

    # 2) Mensajes
    if "message" in update:
        msg = update["message"]
        chat = msg.get("chat", {})
        chat_id = chat.get("id")
        text = (msg.get("text", "") or "").strip()

        if text.startswith("/start"):
            send_message(bot_key, chat_id, f"Hola 👋\nUsa /prueba para ver el error con el botón REPORTAR.\nBot: {bot_display}")
            return {"ok": True}

        if text.startswith("/prueba"):
            err_text = BOTS[bot_key]["default_error_text"]
            err_code = BOTS[bot_key]["default_error_code"]
            kb = {"inline_keyboard": [[{"text": "📩 REPORTAR", "callback_data": f"report:{err_code}"}]]}
            send_message(bot_key, chat_id, err_text, reply_markup=kb)
            return {"ok": True}

        if text.startswith("/getchatid"):
            send_message(bot_key, chat_id, f"chat_id de este chat/grupo: {chat_id}")
            return {"ok": True}

        # /r en el grupo soporte
        if chat_id == SUPPORT_GROUP_ID and (text.startswith("/r") or text.startswith("/r@")):
            # ✅ Solo router procesa /r
            if bot_key != SUPPORT_ROUTER_BOT_KEY:
                return {"ok": True}

            reply_to = msg.get("reply_to_message")
            if not reply_to or not (reply_to.get("text") or ""):
                send_message(bot_key, chat_id, "⚠️ Debes responder (reply) al mensaje del ticket y escribir: /r tu respuesta")
                return {"ok": True}

            replied_text = (reply_to.get("text") or "")
            if TICKET_TAG not in replied_text:
                send_message(bot_key, chat_id, "⚠️ El mensaje respondido no parece un ticket.")
                return {"ok": True}

            raw_ticket_bot = parse_ticket_botkey(replied_text) or ""
            ticket_bot_key = normalize_bot_key(raw_ticket_bot)
            if not ticket_bot_key:
                send_message(bot_key, chat_id, f"⚠️ No pude determinar el bot del ticket. BotKey/BotName leído: {raw_ticket_bot}")
                return {"ok": True}

            m = CHATID_RE.search(replied_text)
            if not m:
                send_message(bot_key, chat_id, "⚠️ No encontré ChatID en el ticket.")
                return {"ok": True}

            client_chat_id = int(m.group(1))

            reply_text = re.sub(r"^/r(@\w+)?\s*", "", text).strip()
            if not reply_text:
                send_message(bot_key, chat_id, "⚠️ Escribe algo después de /r. Ej: /r Ya te ayudamos con el biométrico.")
                return {"ok": True}

            # Enviar al cliente con el bot correcto
            try:
                send_message(ticket_bot_key, client_chat_id, f"🤓 Soporte: {reply_text}")
                send_message(bot_key, chat_id, f"✅ Respuesta enviada al cliente usando {ticket_bot_key}.")
            except Exception as e:
                send_message(bot_key, chat_id, f"❌ No pude enviar al cliente con {ticket_bot_key}: {e}")

            return {"ok": True}

    return {"ok": True}

